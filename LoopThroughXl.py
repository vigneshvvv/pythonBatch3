from openpyxl import load_workbook

workbook = load_workbook("employees.xlsx")

sheet = workbook.active

for row in sheet.iter_rows(values_only=True):
    print(row)